from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler
from openpyxl import load_workbook


TOKEN = '1799218775:AAEBnVCAoRvEfjbBlbuQTotZa88dKABmDZ8'
book = load_workbook('Game List.xlsx')
sheet_1 = book['list1']
gameinfo_page = book['list1']


def main():
    updater = Updater(token=TOKEN)

    dispatcher = updater.dispatcher

    handler = MessageHandler(Filters.all, do_echo)
    start_handler = CommandHandler('start', do_start)
    keybord2_handler = MessageHandler(Filters.text('Вернуться к списку игр'), do_start)
    keybord_handler = MessageHandler(Filters.text, do_something)
    sis_handler = MessageHandler(Filters.text('Минимальные системные требования'), do_write_sis)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(keybord2_handler)
    dispatcher.add_handler(sis_handler)
    dispatcher.add_handler(keybord_handler)
    dispatcher.add_handler(handler)

    updater.start_polling()
    updater.idle()

def do_echo(update, context):


    update.message.reply_text(text='ОШИБКА! Бот еще далек от завершения...')

def do_start(update, context):
    keybord = [
        ['Apex Legends', 'Star Wars Battlefront 2', 'Minecraft Windows 10 edition'],
    ]

    update.message.reply_text(
        text='Ваш список игр:',
        reply_markup=ReplyKeyboardMarkup(keybord, one_time_keyboard=True, resize_keyboard=True)
    )

def do_something(update: Update, context):
    text = update.message.text

    print(gameinfo_page.max_row)
    for row in range(1, gameinfo_page.max_row + 1):
        catch_phrase = gameinfo_page.cell(row=row, column=1).value
        print(f'{row=}')
        print(f'{catch_phrase=}')
        print(f'{text=}')
        if catch_phrase in text:
            context.user_data['Game'] = text
            price = gameinfo_page.cell(row=row, column=5).value
            store = gameinfo_page.cell(row=row, column=4).value
            genre = gameinfo_page.cell(row=row, column=2).value

            update.message.reply_text(f'Жанры: {genre}')
            update.message.reply_text(f'Магазин: {store}')
            update.message.reply_text(f'Цена: {price}')

            keybord2 = [
                ['Минимальные системные требования', 'Вернуться к списку игр']
            ]
            update.message.reply_text(
                text='Выберите действие: ',
                reply_markup = ReplyKeyboardMarkup(keybord2, one_time_keyboard=True, resize_keyboard=True)
            )
            break

    #доделать вторую клавиатуру
def do_write_sis (update: Update, context):
    text = context.user_data['Game']

    for row in range(1, gameinfo_page.max_row + 1):
        catch_phrase = gameinfo_page.cell(row=row, column=1).value
        print(f'{row=}')
        print(f'{catch_phrase=}')
        print(f'{text=}')
        if catch_phrase in text:
            CPU = gameinfo_page.cell(row=row, column=7).value
            RAM = gameinfo_page.cell(row=row, column=8).value
            graphics_card = gameinfo_page.cell(row=row, column=9).value
            operating_system = gameinfo_page.cell(row=row, column=6).value
            video_card_memory = gameinfo_page.cell(row=row, column=11).value
            DirectX = gameinfo_page.cell(row=row, column=12).value
            HDD = gameinfo_page.cell(row=row, column=13).value

            update.message.reply_text(f'Оперативная система: {operating_system}')
            update.message.reply_text(f'Процессор: {CPU}')
            update.message.reply_text(f'ОЗУ: {RAM}')
            update.message.reply_text(f'Видеокарта: {graphics_card}')
            update.message.reply_text(f'Память видеокарты: {video_card_memory}')
            update.message.reply_text(f'DirectX: {DirectX}')
            update.message.reply_text(f'Память на диске: {HDD}')

    return do_start

main()
