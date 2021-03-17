from openpyxl import load_workbook


book = load_workbook('Test.xlsx')
sheet_1 = book['Стикеры']
sticker_page = book['Стикеры']

print(book.worksheets)
print(sticker_page['A1'].value)
for i in range(1, 4):
    print(sticker_page.cell(row=1, column=i).value)

